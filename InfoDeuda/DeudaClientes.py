import os
import math
import re
import numpy as np
from DatosLogin import login
from PIL import Image
#from InfoDeuda.Dias_Deuda_SinCheq import dias_Deuda as dias_Deuda_SinCheques
#from InfoDeuda.Dias_Deuda import dias_Deuda
from Dias_Deuda import dias_Deuda
from Dias_Deuda_SinCheq import dias_Deuda as dias_Deuda_SinCheques
import pandas as pd
import dataframe_image as dfi
import pyodbc #Library to connect to Microsoft SQL Server
from DatosLogin import login #Holds connection data to the SQL Server
import sys
import pathlib
from calendar import monthrange
from datetime import datetime
from datetime import timedelta
import datetime
sys.path.insert(0,str(pathlib.Path(__file__).parent.parent))
import logging
import matplotlib.pyplot as plt
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
        , level=logging.INFO
)
logger = logging.getLogger(__name__)

#################################

tiempoInicio = pd.to_datetime("today")
#########
# Formating display of dataframes with comma separator for numbers
pd.options.display.float_format = "{:20,.0f}".format 
#########

try:
    db_conex = pyodbc.connect(
        "DRIVER={ODBC Driver 17 for SQL Server};\
        SERVER="+login[0]+";\
        DATABASE="+login[1]+";\
        UID="+login[2]+";\
        PWD="+ login[3]
    )
except Exception as e:
    listaErrores = e.args[1].split(".")
    logger.error("\nOcurrió un error al conectar a SQL Server: ")
    for i in listaErrores:
        logger.error(i)
    exit()
###########################################################################
##############################   Recibos AYER  Clientes ###################
###########################################################################
    
recibosAyer = pd.read_sql(
 ''' 

    select SUM(Importe) 'Total Recibos Ayer',NROCLIENTE from RecVenta  
    where FECHASQL < DATEADD(day, 0, CAST(GETDATE() AS date)) 
    and FECHASQL >= DATEADD(day, -1, CAST(GETDATE() AS date))
    and NROCLIENTE > 100000 and NROCLIENTE < 500000 
    and importe > 0


    group by NROCLIENTE

 ''' 
  ,db_conex)

recibosAyer = recibosAyer.convert_dtypes()

recibosAyerEfectivo = pd.read_sql(
 ''' 
    select SUM(r.Importe) 'Total Recibos Negativos',NROCLIENTE from CCRec02 as r left join RecVenta as v
	on r.PTOVTAREC = v.PTOVTA
	and r.NRORECIBO = v.NRORECIBO
	and r.UEN = v.UEN
    where FECHASQL < DATEADD(day, 0, CAST(GETDATE() AS date)) 
    and FECHASQL >= DATEADD(day, -1, CAST(GETDATE() AS date))
    and NROCLIENTE > 100000 and NROCLIENTE < 500000 
    and r.importe < 0

    group by NROCLIENTE
    ''' 
  ,db_conex)
###########################################################################
##############################   Remitos y entregas de Efectivo AYER Clientes ###################
###########################################################################
remitosAyer = pd.read_sql(
 ''' 

select SUM(a.[Remitos Ayer]) as 'Remitos Ayer', NROCLIENTE from 
(select sum(IMPORTE) as 'Remitos Ayer', NROCLIENTE from FacRemDet where FECHASQL < DATEADD(day, 0, CAST(GETDATE() AS date)) 
and FECHASQL >= DATEADD(day, -1, CAST(GETDATE() AS date))
and NROCLIENTE > 100000 and NROCLIENTE < 500000 
and importe > 0
group by NROCLIENTE 

union all

select SUM(-Importe) 'Remitos Ayer',NROCLIENTE
from RecVenta  
where FECHASQL < DATEADD(day, 0, CAST(GETDATE() AS date)) 
and FECHASQL >= DATEADD(day, -1, CAST(GETDATE() AS date))
and NROCLIENTE > 100000 and NROCLIENTE < 500000 
and importe < 0

group by NROCLIENTE) as a group by NROCLIENTE


 ''' 
  ,db_conex)

remitosAyer = remitosAyer.convert_dtypes()

###########################################################################
##############################   CHEQUES Clientes ###########################
###########################################################################

cheques = pd.read_sql(
 ''' 
   DECLARE @ayer DATETIME
	SET @ayer = DATEADD(day, -1, CAST(GETDATE() AS date))
	DECLARE @inicioMesActual DATETIME
	SET @inicioMesActual = DATEADD(month, DATEDIFF(month, 0, @ayer), 0)

	DECLARE @inicioMesAnterior DATETIME
	SET @inicioMesAnterior = DATEADD(M,-1,@inicioMesActual)

	--Divide por la cant de días del mes anterior y multiplica por la cant de días del
	--mes actual
	
	DECLARE @hoy DATETIME
	SET @hoy = DATEADD(DAY, DATEDIFF(DAY, 0, CURRENT_TIMESTAMP), 0)


select NROCLIENTE,
	SUM(r.IMPORTE) CHEQUES
	from CCRec02 as r 
		left join RecVenta as v
		on r.PTOVTAREC = v.PTOVTA
		and r.NRORECIBO = v.NRORECIBO
		and r.UEN = v.UEN
		where  FECHAVTOSQL > @hoy AND FECHASQL <= @hoy  AND FECHASQL >= '2024-01-01'
		and NROCLIENTE > 100000 and NROCLIENTE < 500000 
		and r.MEDIOPAGO in (4,10) GROUP BY NROCLIENTE
 ''' 
  ,db_conex)
cheques = cheques.convert_dtypes()
cheques['NROCLIENTE'] = cheques['NROCLIENTE'].astype(int)

###########################################################################
##############################   Deuda Clientes ###########################
###########################################################################

deuda = pd.read_sql(
 ''' 
    SELECT
            CAST(FRD.[NROCLIENTE] as VARCHAR) as 'NROCLIENTE'
            ,RTRIM(Cli.NOMBRE) as 'NOMBRE'

            ,MIN(CAST(FRD.[FECHASQL] as date)) as 'FECHA_1erRemito'
            ,MAX(CAST(FRD.[FECHASQL] as date)) as 'FECHA_UltRemito'

            ----Días Entre Remitos
            ,IIF(MIN(CAST(FRD.[FECHASQL] as date)) = MAX(CAST(FRD.[FECHASQL] as date))
            	, -1
            	, DATEDIFF(DAY,MAX(CAST(FRD.[FECHASQL] as date)),MIN(CAST(FRD.[FECHASQL] as date)))
            ) as 'Días Entre Remitos'

            --,sum(FRD.[IMPORTE]) as 'ConsumoHistorico'

            ----Consumo Diario
            --,sum(FRD.[IMPORTE])/IIF(MIN(CAST(FRD.[FECHASQL] as date)) = MAX(CAST(FRD.[FECHASQL] as date))
            --	, -1
            --	, DATEDIFF(DAY,MAX(CAST(FRD.[FECHASQL] as date)),MIN(CAST(FRD.[FECHASQL] as date)))
            --) as 'Consumo Diario'

            ,CAST(ROUND(MIN(Cli.SALDOPREPAGO - Cli.SALDOREMIPENDFACTU),0) as int) as 'SALDOCUENTA'

            , DATEDIFF(DAY, MAX(CAST(FRD.[FECHASQL] as date)), CAST(GETDATE() as date)) as 'Días Desde Última Compra'

            , ISNULL(RTRIM(Vend.NOMBREVEND),'') as 'Vendedor'
			, cli.LIMITECREDITO as 'Acuerdo de Descubierto $'
			, sum(FRD.importe) as 'venta en cta/cte'
            FROM [Rumaos].[dbo].[FacRemDet] as FRD with (NOLOCK)
            INNER JOIN dbo.FacCli as Cli with (NOLOCK)
                ON FRD.NROCLIENTE = Cli.NROCLIPRO
            LEFT OUTER JOIN dbo.Vendedores as Vend with (NOLOCK)
	            ON Cli.NROVEND = Vend.NROVEND

            where FRD.NROCLIENTE > '100000'
                and Cli.ListaSaldoCC = 1
                and FECHASQL >= '20140101' and FECHASQL <= CAST(GETDATE() as date)

            group by FRD.NROCLIENTE, Cli.NOMBRE, Vend.NOMBREVEND,cli.LIMITECREDITO
            order by Cli.NOMBRE
 ''' 
  ,db_conex)


deuda = deuda.convert_dtypes()
deuda['NROCLIENTE'] = deuda['NROCLIENTE'].astype(int)
deuda['Cupo Disponible']= deuda['SALDOCUENTA']+deuda['Acuerdo de Descubierto $']
deuda = deuda.merge(cheques,on=['NROCLIENTE'],how='outer')
deuda['CHEQUES'] = deuda['CHEQUES'].fillna(0)
deuda['Saldo Cuenta Con Cheques'] = deuda['SALDOCUENTA']-deuda['CHEQUES']
deuda = deuda[deuda['Saldo Cuenta Con Cheques'] < -1000]
dias_Deuda=dias_Deuda.rename(columns={'Dias de Venta Adeudado':'Dias de Deuda Con Cheques','Cantidad Adeudada':'Cantidad Adeudada Con Cheques'})
deuda=deuda.merge(dias_Deuda,on=['NROCLIENTE'],how='outer')
deuda['Dias de Deuda Con Cheques'] = deuda['Dias de Deuda Con Cheques'].fillna(0)
deuda=deuda.merge(dias_Deuda_SinCheques,on=['NROCLIENTE'],how='outer')
deuda['Dias de Venta Adeudado'] = deuda['Dias de Venta Adeudado'].fillna(0)
deuda['Cantidad Adeudada'] = deuda['Cantidad Adeudada'].fillna(0)
deuda = deuda.merge(remitosAyer,on='NROCLIENTE',how='left')
deuda['Remitos Ayer'] = deuda['Remitos Ayer'].fillna(0)
deuda = deuda.merge(recibosAyer,on='NROCLIENTE',how='left')
deuda['Total Recibos Positivos'] = deuda['Total Recibos Ayer'].fillna(0)

deuda = deuda.merge(recibosAyerEfectivo,on='NROCLIENTE',how='left')
deuda['Total Recibos Negativos'] = deuda['Total Recibos Negativos'].fillna(0)

deuda['Balance Diario']= deuda['Total Recibos Positivos']-deuda['Remitos Ayer']-deuda['Total Recibos Negativos']


def categorizar_deuda(dias_adeuda,ultima_compra):
    if dias_adeuda < 14:
        return '1-Normal'
    elif dias_adeuda < 20:
        return '2-Excedido'
    elif dias_adeuda < 30:
        return '3-Moroso'
    elif dias_adeuda <= 60:
        return '4-Moroso Grave'
    elif dias_adeuda > 60 and ultima_compra > 14:
        return '5-Gestión Judicial'
    else:
        return '4-Moroso Grave'

def categorizar_deuda_Cheques(dias_adeuda_C,ultima_compra_C):
    if dias_adeuda_C < 14:
        return '1-Normal'
    elif dias_adeuda_C < 20:
        return '2-Excedido'
    elif dias_adeuda_C < 30:
        return '3-Moroso'
    elif dias_adeuda_C <= 60:
        return '4-Moroso Grave'
    elif dias_adeuda_C > 60 and ultima_compra_C > 14:
        return '5-Gestión Judicial'
    else:
        return '4-Moroso Grave'

deuda['Dias de Venta Adeudado'] = deuda['Dias de Venta Adeudado'].fillna(1200)
deuda['Días Desde Última Compra'] = deuda['Días Desde Última Compra'].fillna(1200)
deuda['Tipo de deudor'] = deuda.apply(lambda x: categorizar_deuda(x['Dias de Venta Adeudado'], x['Días Desde Última Compra']), axis=1)
deuda['Tipo de deudor Cheques'] = deuda.apply(lambda x: categorizar_deuda(x['Dias de Deuda Con Cheques'], x['Días Desde Última Compra']), axis=1)

# Función para rellenar los valores faltantes según el tipo de dato
deuda['NOMBRE'] = deuda['NOMBRE'].fillna('')
deuda['Cupo Disponible'] = deuda['Cupo Disponible'].fillna(0)
deuda['Acuerdo'] = deuda['Cupo Disponible'].apply(lambda x: 'Deudor Excedido' if x < 0 else 'Deudor No Excedido')
deuda['Acuerdo'] = deuda['Acuerdo'].fillna('')
def calcular_interes(row):
    if row['Dias de Venta Adeudado'] > 15:
        interes_diario = 1.5 / 365  # Tasa de interés diaria (100% anual)
        dias_adeudados = row['Dias de Venta Adeudado'] - 15  # Días de venta adeudados que exceden los 20 días
        saldo_adeudado = -row['Saldo Cuenta Con Cheques']  # Saldo adeudado (en valor absoluto)
        interes_cobrado = saldo_adeudado * interes_diario * dias_adeudados  # Interés cobrado
        return interes_cobrado
    else:
        return 0

# Aplicar la función a cada fila del DataFrame para crear la nueva columna
deuda['Interes por Mora'] = deuda.apply(calcular_interes, axis=1)

deuda2 = deuda.reindex(columns=['SALDOCUENTA','CHEQUES','Saldo Cuenta Con Cheques','Tipo de deudor'])
pagosAyer = deuda.reindex(columns=['NROCLIENTE','NOMBRE','SALDOCUENTA','Saldo Cuenta Con Cheques','Remitos Ayer','Total Recibos Positivos','Total Recibos Negativos','Balance Diario','Tipo de deudor','Acuerdo','Vendedor','Tipo de deudor Cheques'])

deuda = deuda.reindex(columns=['NROCLIENTE','NOMBRE','SALDOCUENTA','Saldo Cuenta Con Cheques','Cantidad Adeudada','Días Desde Última Compra','Acuerdo de Descubierto $','Dias de Venta Adeudado','Dias de Deuda Con Cheques','Cupo Disponible','Interes por Mora','Vendedor','Tipo de deudor','Acuerdo','Tipo de deudor Cheques'])

deuda = deuda.loc[deuda["NOMBRE"] != " ",:]
deuda = deuda.loc[deuda["NOMBRE"] != "",:]
deuda.sort_values(['NOMBRE'])
deuda1 = deuda.reindex(columns=['NROCLIENTE','NOMBRE','SALDOCUENTA','Saldo Cuenta Con Cheques','Cantidad Adeudada','Días Desde Última Compra','Acuerdo de Descubierto $','Dias de Venta Adeudado','Dias de Deuda Con Cheques','Cupo Disponible','Interes por Mora','Vendedor','Tipo de deudor','Acuerdo','Tipo de deudor Cheques'])
normal = deuda1.loc[(deuda1["Tipo de deudor Cheques"] == "1-Normal") | (deuda1["Tipo de deudor Cheques"] == "2-Excedido"),:]
moroso= deuda1.loc[deuda1["Tipo de deudor Cheques"] == "3-Moroso",:]
morosoG= deuda1.loc[deuda1["Tipo de deudor Cheques"] == "4-Moroso Grave",:]
gestionJ= deuda1.loc[deuda1["Tipo de deudor Cheques"] == "5-Gestión Judicial",:]
normal=normal.sort_values(['NOMBRE'])

moroso = moroso.sort_values(by="Saldo Cuenta Con Cheques", ascending=True)
morosoG = morosoG.sort_values(by="Saldo Cuenta Con Cheques", ascending=True)
gestionJ= gestionJ.sort_values(by="Saldo Cuenta Con Cheques", ascending=True)






color_map = {
    '1-Normal': 'green',
    '2-Excedido': 'yellow',
    '3-Moroso': 'orange',
    '4-Moroso Grave': 'red',
    '5-Gestión Judicial': 'maroon',
    'TOTAL':'black',
    ' ':'black'
}
color_map2 = {
    'Deudor No Excedido': 'green',
    'Deudor Excedido': 'red',
    ' ':'black'

}


# Definir un nuevo dataframe con deuda2 (podria haber usado deuda2 directamente)
total = deuda2

# Asegurarse de que los valores de la columna 'CHEQUES' sean negativos
total['CHEQUES'] = total['CHEQUES'] * -1

# Cambiar el nombre de la columna 'CHEQUES' a 'Deuda Garantizada con Cheques'
total = total.rename(columns={'CHEQUES': 'Deuda Garantizada con Cheques'})

# Agrupar por 'Tipo de deudor' y calcular la suma
total = total.groupby(["Tipo de deudor"], as_index=False).sum()

# Calcular la columna 'Participacion' usando 'SALDOCUENTA'
total['Participacion'] = total['SALDOCUENTA'] / total['SALDOCUENTA'].sum()

# Agregar la fila 'colTOTAL' sumando 'SALDOCUENTA' y 'Deuda Garantizada con Cheques', pero no tocar 'Participacion'
total.loc["colTOTAL"] = pd.Series(
    total[["SALDOCUENTA", "Deuda Garantizada con Cheques","Saldo Cuenta Con Cheques"]].sum(),
    index=["SALDOCUENTA", "Deuda Garantizada con Cheques","Saldo Cuenta Con Cheques"]
)

# Rellenar el valor nulo en 'Tipo de deudor' para la fila TOTAL
total.fillna({"Tipo de deudor": "TOTAL"}, inplace=True)

# Rellenar cualquier valor nulo en la columna 'Participacion' con 1
total.fillna({"Participacion": total['SALDOCUENTA'].sum() / total['SALDOCUENTA'].sum()}, inplace=True)


ventas15Dias = pd.read_sql(
 ''' 
   DECLARE @ayer DATETIME
	SET @ayer = DATEADD(day, -1, CAST(GETDATE() AS date))
	DECLARE @inicioMesActual DATETIME
	SET @inicioMesActual = DATEADD(month, DATEDIFF(month, 0, @ayer), 0)

	DECLARE @inicioMesAnterior DATETIME
	SET @inicioMesAnterior = DATEADD(M,-1,@inicioMesActual)

	--Divide por la cant de días del mes anterior y multiplica por la cant de días del
	--mes actual
	
	DECLARE @hoy DATETIME
	SET @hoy = DATEADD(DAY, DATEDIFF(DAY, 0, CURRENT_TIMESTAMP), 0)

	select NROCLIPRO as NROCLIENTE,sum(IMPTOTAL) as 'Ventas Ultimos 15 Dias' from FacRemGen
	where FECHASQL <= @hoy
	and FECHASQL >= DATEADD(DAY,-15,CAST(GETDATE() AS date))
	and NROCLIPRO >= '100000' and NROCLIPRO < '500000'
	group by NROCLIPRO
 ''' 
  ,db_conex)
ventas15Dias=ventas15Dias.convert_dtypes()
deuda = deuda.merge(ventas15Dias,on='NROCLIENTE',how='left')



deuda=deuda.sort_values('Saldo Cuenta Con Cheques', ascending=True)
deuda.loc["colTOTAL"]= pd.Series(
    deuda.sum(numeric_only=True)
    , index=["SALDOCUENTA",'Saldo Cuenta Con Cheques',"Ventas Ultimos 15 Dias","Cantidad Adeudada",'Interes por Mora']
)
deuda.fillna({"NOMBRE":"TOTAL"}, inplace=True)
deuda['Tipo de deudor']=deuda['Tipo de deudor'].fillna(' ')
deuda['Acuerdo']=deuda['Acuerdo'].fillna(' ')

deuda = deuda.reindex(columns=['NROCLIENTE','NOMBRE','SALDOCUENTA','Saldo Cuenta Con Cheques','Cantidad Adeudada','Dias de Venta Adeudado','Dias de Deuda Con Cheques','Días Desde Última Compra','Ventas Ultimos 15 Dias','Acuerdo de Descubierto $','Cupo Disponible','Interes por Mora','Vendedor','Tipo de deudor','Acuerdo','Tipo de deudor Cheques'])

pagosAyer=pagosAyer.sort_values('Saldo Cuenta Con Cheques', ascending=True)
pagosAyer.loc["colTOTAL"]= pd.Series(
    pagosAyer.sum(numeric_only=True)
    , index=["SALDOCUENTA",'Saldo Cuenta Con Cheques','Remitos Ayer','Total Recibos Positivos','Total Recibos Negativos','Balance Diario']
)
pagosAyer.fillna({"NOMBRE":"TOTAL"}, inplace=True)
pagosAyer['Tipo de deudor']=deuda['Tipo de deudor'].fillna('')
pagosAyer['Acuerdo']=deuda['Acuerdo'].fillna('')


def _estiladorVtaTituloD(df,list_Col_Perc,list_Col_Numpes, list_Col_Num,listaporcentaje, titulo):
    """
This function will return a styled dataframe that must be assign to a variable.
ARGS:
    df: Dataframe that will be styled.
    list_Col_Num: List of numeric columns that will be formatted with
    zero decimals and thousand separator.
    list_Col_Perc: List of numeric columns that will be formatted 
    as percentage.
    titulo: String for the table caption.
    """
    excluded_rows = df.index[-1:]
    resultado = df.style \
        .format("$ {0:,.0f}", subset=list_Col_Numpes) \
        .format("{0:,.0f}", subset=list_Col_Num) \
        .format("{:,.2%}", subset=listaporcentaje) \
        .hide(axis=0) \
        .set_caption(
            titulo
            + "<br>"
            + ((tiempoInicio-pd.to_timedelta(0,"days")).strftime("%d/%m/%y"))
            + "<br>") \
        .set_properties(subset= list_Col_Perc + list_Col_Num +list_Col_Numpes+listaporcentaje
            , **{"text-align": "center", "width": "100px"}) \
        .set_properties(border= "2px solid black") \
        .set_table_styles([
            {"selector": "caption", 
                "props": [
                    ("font-size", "20px")
                    ,("text-align", "center")
                ]
            }
            , {"selector": "th", 
                "props": [
                    ("text-align", "center")
                    ,("background-color","black")
                    ,("color","white")
                ]
            }
        ]) \
        .apply(lambda x: ["background: black" if x.name == "colTOTAL" 
            else "" for i in x]
            , axis=1) \
        .apply(lambda x: ["color: white" if x.name == "colTOTAL" 
            else "" for i in x]
            , axis=1)\
        .map(
            lambda x: f'background-color: {color_map.get(x, "")}' if pd.notna(x) and x not in excluded_rows else '', 
            subset=['Tipo de deudor']) \
        .map(
            lambda x: f'background-color: {color_map.get(x, "")}' if pd.notna(x) and x not in excluded_rows else '', 
            subset=['Tipo de deudor Cheques']) \
        .map(
            lambda x: f'background-color: {color_map2.get(x, "")}' if pd.notna(x) and x not in excluded_rows else '', 
            subset=['Acuerdo'])
    return resultado


######### LE DOY FORMATO AL DATAFRAME
def _estiladorVtaTituloDTotal(df,list_Col_Numpes,listaporcentaje, titulo):
    """
This function will return a styled dataframe that must be assign to a variable.
ARGS:
    df: Dataframe that will be styled.
    list_Col_Num: List of numeric columns that will be formatted with
    zero decimals and thousand separator.
    list_Col_Perc: List of numeric columns that will be formatted 
    as percentage.
    titulo: String for the table caption.
    """
    resultado = df.style \
        .format("$ {0:,.0f}", subset=list_Col_Numpes) \
        .format("{:,.2%}", subset=listaporcentaje) \
        .hide(axis=0) \
        .set_caption(
            titulo
            + "<br>"
            + ((tiempoInicio - pd.to_timedelta(0, "days")).strftime("%d/%m/%y"))
            + "<br>") \
        .set_properties(
            subset=list_Col_Numpes + listaporcentaje,
            **{"text-align": "center", "width": "150px"}) \
        .set_properties(border="2px solid black") \
        .set_table_styles([
            {"selector": "caption",
            "props": [
                ("font-size", "20px"),
                ("text-align", "center")
         ]
        },
        {"selector": "th",
         "props": [
             ("text-align", "center"),
             ("background-color", "black"),
             ("color", "white")
         ]
        }
    ]) \
        .map(lambda x: f'background-color: {color_map.get(x, "")}', subset=['Tipo de deudor']) \
        .apply(lambda x: ["background: black" if x.name == "colTOTAL" else "" for i in x], axis=1) \
        .apply(lambda x: ["color: white" if x.name == "colTOTAL" else "" for i in x], axis=1)
     
    return resultado


#  columnas sin decimales
numCols1=[]
numCols = [ 'NROCLIENTE','Días Desde Última Compra','Dias de Venta Adeudado','Dias de Deuda Con Cheques','Cantidad Adeudada']
colpesos=[ 'SALDOCUENTA','Saldo Cuenta Con Cheques','Acuerdo de Descubierto $','Cupo Disponible','Interes por Mora']
coldeuda=['SALDOCUENTA','Deuda Garantizada con Cheques','Saldo Cuenta Con Cheques']
colpesos1=[ 'SALDOCUENTA','Saldo Cuenta Con Cheques','Total Recibos Positivos','Remitos Ayer','Total Recibos Negativos','Balance Diario']
# Columnas Porcentaje
percColsstr = ['Tipo de deudor','Acuerdo','Tipo de deudor Cheques']
percColsstr1 = ['Tipo de deudor']
percCol = ['Participacion']



deuda = _estiladorVtaTituloD(deuda,percColsstr,colpesos,numCols,numCols1, "deuda")

pagosAyer = _estiladorVtaTituloD(pagosAyer,percColsstr,colpesos1,numCols1,numCols1, "Remitos y Recibos Ayer")

moroso = _estiladorVtaTituloD(moroso,percColsstr,colpesos,numCols,numCols1, "deuda")

morosoG = _estiladorVtaTituloD(morosoG,percColsstr,colpesos,numCols,numCols1, "deuda")

gestionJ = _estiladorVtaTituloD(gestionJ,percColsstr,colpesos,numCols,numCols1, "deuda")


total = _estiladorVtaTituloDTotal(total,coldeuda,percCol, "deuda")


### APLICO EL FORMATO A LA TABLA
ubicacion = str(pathlib.Path(__file__).parent)+"\\"
nombreN = "Deudores_Comerciales.png"
nombreM= "Deudores_Morosos.png"
nombreMG = "Deudores_MorososG.png"
nombreGJ = "Deudores_GestionJ.png"
nombreT='Deuda_Comercial.png'

def df_to_image(df, ubicacion, nombre):
    """
    Esta función usa las biblioteca "dataframe_Image as dfi" y "os" para 
    generar un archivo .png de un dataframe. Si el archivo ya existe, este será
    reemplazado por el nuevo archivo.

    Args:
        df: dataframe a convertir
         ubicacion: ubicacion local donde se quiere grabar el archivo
          nombre: nombre del archivo incluyendo extensión .png (ej: "hello.png")

    """
    if os.path.exists(ubicacion+nombre):
        os.remove(ubicacion+nombre)
        dfi.export(df, ubicacion+nombre,max_rows=-1)
    else:
        dfi.export(df, ubicacion+nombre,max_rows=-1)

df_to_image(deuda, ubicacion, nombreN)
df_to_image(moroso, ubicacion, nombreM)
df_to_image(morosoG, ubicacion, nombreMG)
df_to_image(gestionJ, ubicacion, nombreGJ) 
df_to_image(total, ubicacion, nombreT)


imgRGBA = Image.open(ubicacion+nombreN)
if len(imgRGBA.split()) == 4:
    # Si tiene un canal alfa, crear una nueva imagen RGB y pegar la imagen RGBA con el canal alfa como máscara
    imRGB = Image.new('RGB', imgRGBA.size, (255, 255, 255))
    imRGB.paste(imgRGBA, mask=imgRGBA.split()[3])
else:
    # Si no tiene un canal alfa, simplemente convertir la imagen a RGB
    imRGB = imgRGBA.convert('RGB')
nombrePDF = "Deudores_Comerciales.pdf"

# Saving has a PDF and avoiding permission error
if os.path.exists(ubicacion + nombrePDF):
    os.remove(ubicacion + nombrePDF)
    imRGB.save(ubicacion + nombrePDF, "PDF", resolution=90.0, save_all=True)
else:
    imRGB.save(ubicacion + nombrePDF, "PDF", resolution=90.0, save_all=True)

### EXCEL
ubicacionExcel = str(pathlib.Path(__file__).parent)+"\\"

nombreExcel = "Clientes_Deudores.xlsx"
nombreExcelRecibos = "Remitos_Recibos_Ayer.xlsx"
### IMPRIMO LA Excel 

def df_to_Excel(df, ubicacion, nombre):
    """
    Esta función usa las biblioteca "dataframe_Image as dfi" y "os" para 
    generar un archivo .png de un dataframe. Si el archivo ya existe, este será
    reemplazado por el nuevo archivo.

    Args:
        df: dataframe a convertir
         ubicacion: ubicacion local donde se quiere grabar el archivo
          nombre: nombre del archivo incluyendo extensión .png (ej: "hello.png")

    """
    if os.path.exists(ubicacion+nombre):
        os.remove(ubicacion+nombre)
        df.to_excel(ubicacion+nombre)
    else:
        df.to_excel(ubicacion+nombre)

df_to_Excel(deuda, ubicacionExcel, nombreExcel)
df_to_Excel(pagosAyer, ubicacionExcel, nombreExcelRecibos)


import openpyxl
from openpyxl.styles import PatternFill, Font

############################## Editar Excel ### 
# Ruta del archivo Excel
#archivo_ruta = 'C:/Informes/InfoDeuda/Clientes_Deudores.xlsx'
archivo_ruta = 'C:/Users/mespinosa/AppData/Local/Microsoft/WindowsApps/Informes/InfoDeuda/Clientes_Deudores.xlsx'

# Abrir el archivo Excel
archivo_excel = openpyxl.load_workbook(archivo_ruta)

# Obtener la hoja de cálculo activa
hoja_activa = archivo_excel.active

# Agrandar las columnas B a L
hoja_activa.column_dimensions.group('B', 'L', hidden=False)

# Obtener el número de la última fila escrita
ultima_fila = hoja_activa.max_row

# Pintar la última fila desde la columna B hasta la K
fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
for col in range(2, 12):  # Columnas B a K
    celda = hoja_activa.cell(row=ultima_fila, column=col)
    celda.fill = fill
    celda.font = Font(color="FFFFFF")  # Blanco

# Guardar los cambios en el archivo Excel
archivo_excel.save(archivo_ruta)

# Cerrar el archivo Excel
archivo_excel.close()

############################## Editar 2do Excel ### 

# Cargar el archivo Excel
archivo_excel2 = openpyxl.load_workbook(str(pathlib.Path(__file__).parent)+'/Remitos_Recibos_Ayer.xlsx')

# Obtener la hoja de cálculo activa
hoja_activa2 = archivo_excel2.active

# Agrandar las columnas A a H
hoja_activa2.column_dimensions.group('B', 'L', hidden=False)

# Obtener el número de la última fila escrita
ultima_fila2 = hoja_activa2.max_row

# Pintar la última fila desde la columna B hasta la K
fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
for col in range(2, 12):  # Columnas B a K
    celda = hoja_activa2.cell(row=ultima_fila2, column=col)
    celda.fill = fill
    celda.font = Font(color="FFFFFF")  # Blanco

# Guardar los cambios en el archivo Excel
archivo_excel2.save(str(pathlib.Path(__file__).parent)+'/Remitos_Recibos_Ayer.xlsx')

# Cerrar el archivo
archivo_excel2.close()


#############
# Timer
tiempoFinal = pd.to_datetime("today")
logger.info(
    "\nInfo Volumen de Ventas"
    + "\nTiempo de Ejecucion Total: "
    + str(tiempoFinal-tiempoInicio)
)
